import openpyxl, os
from openpyxl.styles import Font, Alignment

def update_table(task, ext_mode, goal_prob, step_size, num_iter, time, cost, coverage=None):
    if not os.path.exists('Output/Results.xlsx'):
        wb = openpyxl.Workbook()
        ws1 =  wb.active
        ws1.title = 'MP'
        ws2 = wb.create_sheet('IP')
        for idx, title in enumerate(['Ext. Mode','Goal Bias','Step Size','Avg. Cost','Avg. Time [sec]','Avg. Num. of Iterations']):
            ws1.cell(1,idx+1,value = title)
            ws1.cell(1,idx+1).font = Font(bold=True)
            ws1.cell(1,idx+1).alignment = Alignment(horizontal='center')
        for idx, title in enumerate(['Ext. Mode','Goal Bias','Step Size','Coverage','Avg. Cost','Avg. Time [sec]','Avg. Num. of Iterations']):
            ws2.cell(1,idx+1,value = title)
            ws2.cell(1,idx+1).font = Font(bold=True)
            ws2.cell(1,idx+1).alignment = Alignment(horizontal='center')

        ws1.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 12
        ws1.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
        ws1.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 10
        ws1.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
        ws1.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 15
        ws1.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 22
        
        ws2.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 12
        ws2.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
        ws2.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 10
        ws2.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
        ws2.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 10
        ws2.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 15
        ws2.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 22
        
        wb.save(filename='Output/Results.xlsx')
    else:
        wb = openpyxl.load_workbook('Output/Results.xlsx')

    if task=='mp':
        ws = wb['MP']
        i = 2
        while ws.cell(i,1).value is not None:
           i += 1
        ws.cell(i, 1, value = ext_mode)
        ws.cell(i, 2, value = goal_prob)
        if ext_mode=='E2':
            ws.cell(i, 3, value = step_size)
        else:
            ws.cell(i, 3, value = '-')
        ws.cell(i, 4, value = f"{cost:0.3f}")
        ws.cell(i, 5, value = f"{time:0.3f}")
        ws.cell(i, 6, value = f"{num_iter:0.0f}")
        for j in range(1,7):
            ws.cell(i,j).alignment = Alignment(horizontal='center')
    
    elif task=='ip':
        ws = wb['IP']
        i = 2
        while ws.cell(i,1).value is not None:
           i += 1
        ws.cell(i, 1, value = ext_mode)
        ws.cell(i, 2, value = goal_prob)
        if ext_mode=='E2':
            ws.cell(i, 3, value = step_size)
        else:
            ws.cell(i, 3, value = '-')
        ws.cell(i, 4, value = coverage)
        ws.cell(i, 5, value = f"{cost:0.3f}")
        ws.cell(i, 6, value = f"{time:0.3f}")
        ws.cell(i, 7, value = f"{num_iter:0.0f}")
        for j in range(1,8):
            ws.cell(i,j).alignment = Alignment(horizontal='center')
    
    wb.save(filename='Output/Results.xlsx')
    return